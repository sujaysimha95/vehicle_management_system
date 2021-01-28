"""
Created on Sat Dec 26 20:27:19 2020

@author: $ujay $imha
"""

# module imports
from flask import Flask, render_template, redirect, url_for, request
from mysql import connector
from pymsgbox import alert
from tzlocal import get_localzone
import openpyxl
import platform
# import logging
import simple_log
import arrow

# defining the app and setting dev environment
app = Flask(__name__, static_folder='templates/static')
app.config['ENV'] = 'development'

# logging.basicConfig(filename='vehicle service management.log', level=0)
# adding log file
log = simple_log.get_log('vehicle service management')
simple_log.add_file('vehicle service management', 'vehicle service management.log')
# opening log file and adding timestamp
log_f = open('vehicle service management.log', 'a')
log_f.write('\n')
log_f.write(f'timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}; timestamp(sys time): '
            f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}')
log_f.close()

# connecting to mysql database
log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
         f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; Connecting to database...')
conn = connector.connect(host='localhost', user='root', password='mysql@0123456789',
                         database='vehicle_management_system')
cursor = conn.cursor(buffered=True)
log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
         f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; Connected to database!')

# retrieving names, accounts, passwords and numbers from the mysql database
names = []
accounts = []
passwords = []
numbers = []
log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
         f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; Fetching data from table '
         '`vehicle_management_system`.`users`...')
cursor.execute("SELECT name FROM users;")
log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
         f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
         f'Done!')
for i in cursor.fetchall():
    names.append(i[0])
cursor.execute("SELECT account FROM users;")
for i in cursor.fetchall():
    accounts.append(i[0])
cursor.execute("SELECT password FROM users;")
for i in cursor.fetchall():
    passwords.append(i[0])
cursor.execute("SELECT phone_number FROM users;")
for i in cursor.fetchall():
    numbers.append(i[0])


# user functionalities start from here


# main sign-up page
@app.route('/sign-up')
def sign_up():
    return render_template(r'sign-up page.html')


# sign up page for user
# noinspection PyShadowingNames
@app.route('/sign-up/user', methods=["POST", "GET"])
def sign_up_as_user():
    active_users = []
    cursor.execute("SELECT * FROM users WHERE currently_active='true';")
    for i in cursor.fetchall():
        active_users.append(i)
    if len(active_users) == 0:
        if request.method == "POST":
            log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                     f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                     f'Fetching data from html form')
            usr = request.form["username"]
            acc = request.form["account"]
            passw = request.form["password"]
            num = request.form["number"]
            log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                     f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                     f'Done!')

            if len(num) != 10:
                alert(text='The length of the number must be 10 characters long!', title='Warning', button='OK')
                return redirect(url_for('sign_up_as_user'))
            if len(num) == 10:
                log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                         f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; Saving '
                         f'credentials into table `vehicle_management_system`.`users`...')
                cursor.execute(f"INSERT INTO users(name, account, password, phone_number, currently_active) "
                               f"VALUES('{usr}', '{acc}', '{passw}', '{num}', "
                               f"'true');")
                conn.commit()
                log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                         f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                         f'Sign-up successful!')
                log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                         f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                         f'Data retrieved:\nusername: {usr}\naccount: {acc}\npassword: {passw}\nphone number: '
                         f'{num};')
                return redirect(url_for('dashboard'))
        else:
            return render_template('sign-up page(user).html')

    if len(active_users) > 0:
        alert(text='You have already been logged in. You will now be directed to the dashboard', title='Note')
        return redirect(url_for('dashboard'))


# login page for user
# noinspection PyShadowingNames
@app.route('/login/user', methods=['GET', 'POST'])
def login_as_user():
    # making 4 variables global
    global names, accounts, passwords, numbers
    active_users = []
    # process for signing up
    cursor.execute("SELECT * FROM users WHERE currently_active='true';")
    for i in cursor.fetchall():
        active_users.append(i)
    if len(active_users) == 0:
        if request.method == 'POST':
            log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                     f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                     f'Fetching data from form')
            # getting username, account and password from html form
            usr = request.form["username"]
            acc = request.form["acc"]
            passw = request.form["passw"]
            log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                     f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                     f'Done!')
            log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                     f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                     f'Data retrieved:\nusername: {usr}\naccount: {acc}\npassword: {passw};')
            # number = request.form["number"]
            if usr not in names:
                alert(text='Username not found. Try again', title='Warning')
                return redirect(url_for('login_as_user'))
            if acc not in accounts:
                alert(text='Account not found. Try again', title='Warning')
                return redirect(url_for('login_as_user'))
            if passw not in passwords:
                alert(text='Password not found. Try again', title='Warning')
                return redirect(url_for('login_as_user'))
            """if number not in numbers:
                alert(text='Phone number not found. Try again', title='Warning')
                return redirect(url_for('login_as_user'))"""
            if usr in names and acc in accounts and passw in passwords:
                # if number in numbers:
                log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                         f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                         f'Logging in...')
                cursor.execute(f"UPDATE users SET currently_active='true' WHERE name='{usr}' AND account='{acc}'"
                               f" AND password='{passw}';")
                conn.commit()
                log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                         f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                         f'login successful!')
                active_users = []
                cursor.execute("SELECT * FROM users WHERE currently_active='true';")
                for i in cursor.fetchall():
                    active_users.append(i)
                if len(active_users) == 0:
                    alert(text='Your username, account and password don\'t match. Try again.',
                          title='Warning')
                    return redirect(url_for('login_as_user'))
                return redirect(url_for('dashboard'))
        else:
            return render_template('login-page(user).html')
    if len(active_users) > 0:
        alert(text='You are already logged in. Please log out to access this page', title='Note')
        return redirect(url_for('dashboard'))


# logout page
@app.route('/logout')
def logout():
    cursor.execute(f"update users set currently_active='false';")
    conn.commit()
    log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
             f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
             f'Logout successful!')
    return redirect(url_for('sign_up'))


# user dashboard
@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')


# user service-request
# noinspection PyShadowingNames
@app.route('/service-request')
def service_request():
    # displays the service-requests the user has submitted
    active_user = []
    cursor.execute("SELECT name FROM users WHERE currently_active='true';")
    for i in cursor.fetchall():
        active_user.append(i[0])
    log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
             f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
             f'Fetching data from table `vehicle_management_system`.`service_request`...')
    cursor.execute(f"SELECT vehicle_category, vehicle_name, vehicle_model, vehicle_brand, "
                   f"vehicle_registration_number, service_date, service_time, service_type, delivery_date, "
                   f"delivery_time,delivery_address, bill, status FROM service_request WHERE "
                   f"submitted_user='{active_user[0]}';")
    log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
             f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
             f'Done!')
    return render_template('service-request.html', service_requests=cursor.fetchall())


# new service-request
# noinspection PyShadowingNames
@app.route('/service-request/new', methods=['GET', 'POST'])
def new_service_request():
    if request.method == 'POST':
        active_user = ''
        cursor.execute("SELECT name FROM users WHERE currently_active='true';")
        for i in cursor.fetchall():
            active_user = i[0]
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Getting data from form...')
        # getting stuff from HTML form
        vehicle_catg = request.form.get("vehicle_category")
        vehicle_name = request.form.get("vehicle_name")
        vehicle_model = request.form.get("vehicle_model")
        vehicle_brand = request.form.get("vehicle_brand")
        vehicle_reg_num = request.form.get("vehicle_registration_number")
        service_date = request.form.get("service_date")
        service_time = request.form.get("service_time")
        service_type = request.form.get("service")
        # delivery_date = request.form.get("delivery_date")
        # delivery_time = request.form.get("delivery_time")
        delivery_address = request.form.get("delivery_address")
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Done!')
        bill = 0
        if service_type.lower() == 'general service':
            bill = 500
        if service_type.lower() == 'special service':
            bill = 1000
        # inserting the service-request into table
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Saving data into table `vehicle_management_system`.`service_request`...')
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Data retrieved:\nVehicle category: {vehicle_catg}\nVehicle name: {vehicle_name}\n'
                 f'Vehicle model: {vehicle_model}\nVehicle brand: {vehicle_brand}\n'
                 f'Vehicle registration number: {vehicle_reg_num}\nService type: {service_type}\n'
                 f'Service date: {service_date}\nService time: {service_time}\nDelivery address: '
                 f'{delivery_address};')
        cursor.execute(f"INSERT INTO service_request(submitted_user, vehicle_category, vehicle_name, vehicle_model, "
                       f"vehicle_brand, vehicle_registration_number, service_date, service_time, service_type, "
                       f"delivery_address, bill, status) VALUES('{active_user}', "
                       f"'{vehicle_catg}', '{vehicle_name}', '{vehicle_model}', '{vehicle_brand}', '{vehicle_reg_num}',"
                       f" '{service_date}', '{service_time}','{service_type}', "
                       f"'{delivery_address}', '{bill}', 'pending');")
        """cursor.execute(f"INSERT INTO service_request(submitted_user, vehicle_category, vehicle_name, vehicle_model, "
                       f"vehicle_brand, vehicle_registration_number, service_type, "
                       f"delivery_date, delivery_time, delivery_address, bill, status) VALUES('{active_user}', "
                       f"'{vehicle_catg}', '{vehicle_name}', '{vehicle_model}', '{vehicle_brand}', '{vehicle_reg_num}',"
                       f" '{service_type}', '{delivery_date}', '{delivery_time}', '{delivery_address}', '{bill}', "
                       f"'pending');")"""
        conn.commit()
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Done!')
        return redirect(url_for('service_request'))
    else:
        return render_template('new-service-request.html')


# user profile
# noinspection PyShadowingNames
@app.route('/profile')
def profile():
    currently_active = []
    log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
             f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
             f'Fetching data from table `vehicle_management_system`.`users`')
    cursor.execute("SELECT name, account, phone_number FROM users WHERE currently_active='true';")
    log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
             f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
             f'Done!')
    for i in cursor.fetchall():
        currently_active.append(i)
    usr = currently_active[0][0]
    acc = currently_active[0][1]
    num = currently_active[0][2]
    return render_template('profile.html', name=usr, acc=acc, num=num)


# updating user profile
# noinspection PyShadowingNames
@app.route('/profile/update', methods=["GET", "POST"])
def update_profile():
    active_users = []
    cursor.execute("SELECT * FROM users WHERE currently_active='true';")
    for i in cursor.fetchall():
        active_users.append(i)
    if request.method == "POST":
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Fetching data from form...')
        usr = request.form.get("username")
        acc = request.form.get("acc")
        num = request.form.get("num")
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Done!')
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Data retrieved:\nusername: {usr}\naccount: {acc}\nphone number: {num};')

        if usr == '':
            usr = active_users[0][1]
        if acc == '':
            acc = active_users[0][2]
        if num == '':
            num = active_users[0][4]
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Updating user credentials...')
        cursor.execute(f"UPDATE users SET name='{usr}', account='{acc}', phone_number='{num}' WHERE "
                       f"id={active_users[0][0]};")
        conn.commit()
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Done!')
        return redirect(url_for('profile'))
    else:
        return render_template('update-profile.html')


# updating user password
# noinspection PyShadowingNames
@app.route('/profile/update/password', methods=['GET', 'POST'])
def change_password():
    active_users = []
    cursor.execute("SELECT * FROM users WHERE currently_active='true';")
    for i in cursor.fetchall():
        active_users.append(i)
    if request.method == 'POST':
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Fetching data from form...')
        current_passw = request.form.get('current_password')
        new_passw = request.form.get('new_password')
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Data retrieved:\nCurrent Password: {current_passw}\nNew Password: {new_passw};')
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Done!')
        if current_passw != active_users[0][3]:
            alert(text='The password you entered does not match your current password. Try again', title='Warning')
            return redirect(url_for('change_password'))
        if current_passw == active_users[0][3]:
            log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                     f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                     f'Updating password...')
            cursor.execute(f"UPDATE users SET password='{new_passw}' WHERE name='{active_users[0][1]}' AND "
                           f"account='{active_users[0][2]}' AND phone_number='{active_users[0][4]}';")
            conn.commit()
            log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                     f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                     f'Password successfully updated!')
            return redirect(url_for('profile'))
    else:
        return render_template('update-password.html')


# admin functionalities start from here


# admin logging page
# noinspection PyShadowingNames
@app.route('/login/admin', methods=['GET', 'POST'])
def login_as_admin():
    active_users = []
    cursor.execute("SELECT * FROM users WHERE currently_active='true';")
    for i in cursor.fetchall():
        active_users.append(i)
    if len(active_users) == 0:
        admin_cred = []
        cursor.execute("SELECT name, password FROM users;")
        for i in cursor.fetchall():
            admin_cred.append(i)
        if request.method == 'POST':
            log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                     f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                     f'Fetching data from form...')
            usr = request.form.get("username")
            passw = request.form.get("passw")
            log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                     f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                     f'Done!')
            log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                     f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                     f'Data retrieved:\nusername: {usr}\npassword: {passw};')
            if usr != admin_cred[0][0]:
                alert(text='Username does not match! Try again', title='Warning')
                return redirect(url_for('login_as_admin'))
            if passw != admin_cred[0][1]:
                alert(text='Password Does not match! Try again', title='Warning')
                return redirect(url_for('login_as_admin'))
            if usr == admin_cred[0][0] and passw == admin_cred[0][1]:
                log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                         f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                         f'Logging in...')
                cursor.execute("UPDATE users SET currently_active='true' WHERE id=1;")
                conn.commit()
                log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                         f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                         f'Successfully logged in!')
                return redirect(url_for('admin_dashboard'))
        else:
            return render_template('login-page(admin).html')
    if len(active_users) > 0:
        alert(text='You are already logged in. Please log out to access this page', title='Warning')
        if active_users[0][1] == 'admin':
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('dashboard'))


# admin dashboard
@app.route('/admin/dashboard')
def admin_dashboard():
    return render_template('dashboard(admin).html')


# admin service request
@app.route('/admin/service-request', methods=['GET', 'POST'])
def admin_service_request():
    # showing the service requests submitted by all users
    count = range(1, 10)
    log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
             f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
             f'Fetching data from `vehicle_management_system`.`users`...')
    cursor.execute("SELECT * FROM service_request;")
    log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
             f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
             f'Done!')
    return render_template('service-request(admin).html', service_requests=cursor.fetchall(), count=count)


# admin approve page (where he/she can approve a service request)
# noinspection PyShadowingNames
@app.route('/admin/approve', methods=['GET', 'POST'])
def approve():
    # getting id from HTML form
    if request.method == 'POST':
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Fetching data from form...')
        id = request.form.get("id")
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Done!')
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Data retrieved:\nid: {id};')
        # updating database to set the service request to 'approved'
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Approving service request...')
        cursor.execute(f"UPDATE service_request SET status='approved' WHERE id={id};")
        conn.commit()
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Done!')
        return redirect(url_for('admin_service_request'))
    else:
        return render_template('approve.html')


# admin decline page (where he/she can decline a service request)
# noinspection PyShadowingNames
@app.route('/admin/decline', methods=['GET', 'POST'])
def decline():
    # getting id from HTML form
    if request.method == 'POST':
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Getting data from form...')
        id = request.form.get("id")
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Done!')
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Data retrieved:\nid: {id};')
        # updating database to set a service request to 'declined'
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Declining service request...')
        cursor.execute(f"UPDATE service_request SET status='declined' WHERE id={id};")
        conn.commit()
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Done!')
        return redirect(url_for('admin_service_request'))
    else:
        return render_template('decline.html')


# noinspection PyShadowingNames
@app.route('/admin/delivery-date-time', methods=['GET', 'POST'])
def service_date():
    if request.method == 'POST':
        id = request.form.get("id")
        delivery_date = request.form.get("delivery-date")
        delivery_time = request.form.get("delivery-time")
        cursor.execute(f"UPDATE service_request SET delivery_date='{delivery_date}', "
                       f"delivery_time='{delivery_time}' WHERE id={id};")
        conn.commit()
        return redirect(url_for('admin_service_request'))
    else:
        return render_template('delivery-date-time.html')


# admin print bill (where he/she can save a bill to an xlsx file)
# noinspection PyShadowingNames
@app.route('/admin/print-bill', methods=['GET', 'POST'])
def admin_print_bill():
    # getting id from HTML form
    if request.method == 'POST':
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Getting id and path from form...')
        id = request.form.get("id")
        path = request.form.get("path")
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Done!')
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Data retrieved:\nid: {id}\npath: {path};')
        if id == '' and path == '':
            alert(text='ID and path are empty!', title='Warning')
            return redirect(url_for('admin_print_bill'))
        if id == '':
            alert(text='ID is empty!', title='Warning')
            return redirect(url_for('admin_print_bill'))
        if path == '':
            alert(text='path is empty!', title='Warning')
            return redirect(url_for('admin_print_bill'))

        # checking if id isn't all. if no, get id and save the service request with that id.
        if id.lower() != 'all':
            # retrieve the service request
            log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                     f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                     f'Saving service request with id of {id}...')
            data = []
            cursor.execute(f"SELECT * FROM service_request WHERE id={id};")
            # create the path to save the xlsx file
            for i in cursor.fetchall():
                service_request.append(i)
            if r'\bill.xlsx' not in path or '/bill.xlsx' not in path:
                if platform.system() == 'Windows':
                    path += r'\bill.xlsx'
                else:
                    path += '/bill.xlsx'
            # creating a new xlsx file
            wb = openpyxl.Workbook()
            wb.save(path)
            # open the xlsx file we saved in the previous file
            wb = openpyxl.load_workbook(path)
            # select the active sheet
            sheet = wb.active
            # save the data
            for row in data:
                sheet.append(row)
            wb.save(path)
            log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                     f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                     f'Done!')
            return redirect(url_for('admin_print_bill'))

        # checking if id is all. if yes, save every service request
        if id.lower() == 'all':
            # retrieve all service requests
            log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                     f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                     f'Saving all service requests...')
            cursor.execute(f"SELECT * FROM service_request;")
            data = []
            for i in cursor.fetchall():
                data.append(i)
            # create the path to save the xlsx file
            if r'\bill.xlsx' not in path or '/bill.xlsx' not in path:
                if platform.system() == 'Windows':
                    path += r'\bill.xlsx'
                else:
                    path += '/bill.xlsx'
            # create a new xlsx file
            wb = openpyxl.Workbook()
            wb.save(path)
            # open the xlsx file we saved in the previous file
            wb = openpyxl.load_workbook(path)
            # select the active sheet
            sheet = wb.active
            # save the data
            for row in data:
                sheet.append(row)
            wb.save(path)
            log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                     f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                     f'Done!')
            return redirect(url_for('admin_print_bill'))

    else:
        return render_template('print-bill.html')


# admin profile
# noinspection PyShadowingNames
@app.route('/admin/profile')
def admin_profile():
    currently_active = []
    log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
             f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
             f'Fetching data from `vehicle_management_system`.`users`...')
    cursor.execute("SELECT name, account, phone_number FROM users WHERE currently_active='true';")
    log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
             f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
             f'Done!')
    for i in cursor.fetchall():
        currently_active.append(i)
    usr = currently_active[0][0]
    acc = currently_active[0][1]
    num = currently_active[0][2]
    return render_template('admin-profile.html', name=usr, acc=acc, num=num)


# admin update profile
# noinspection PyShadowingNames
@app.route('/admin/profile/update', methods=['GET', 'POST'])
def admin_update_profile():
    active_users = []
    cursor.execute("SELECT * FROM users WHERE currently_active='true';")
    for i in cursor.fetchall():
        active_users.append(i)
    if request.method == "POST":
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Fetching data from form...')
        usr = request.form.get("username")
        acc = request.form.get("acc")
        num = request.form.get("num")
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Done!')
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Data retrieved: \nusername: {usr}\naccount: {acc}\nphone number: {num};')
        if usr == '':
            usr = active_users[0][1]
        if acc == '':
            acc = active_users[0][2]
        if num == '':
            num = active_users[0][4]
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Updating user credentials')
        cursor.execute(f"UPDATE users SET name='{usr}', account='{acc}', phone_number='{num}' WHERE "
                       f"id={active_users[0][0]};")
        conn.commit()
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Done!')
        return redirect(url_for('admin_profile'))
    else:
        return render_template('admin-update-profile.html')


# update password
# noinspection PyShadowingNames
@app.route('/admin/profile/update/password', methods=['GET', 'POST'])
def admin_change_password():
    active_users = []
    log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
             f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; ' 
             f'Fetching data from `vehicle_management_system`.`users`...')
    cursor.execute("SELECT * FROM users WHERE currently_active='true';")
    log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
             f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
             f'Done!')
    for i in cursor.fetchall():
        active_users.append(i)
    if request.method == 'POST':
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Fetching data from form...')
        current_passw = request.form.get('current_password')
        new_passw = request.form.get('new_password')
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Done!')
        log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                 f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                 f'Data retrieved:\nCurrent password: {current_passw}\nNew password: {new_passw};')
        if current_passw != active_users[0][3]:
            alert(text='The password you entered does not match your current password. Try again', title='Warning')
            return redirect(url_for('change_password'))
        if current_passw == active_users[0][3]:
            log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                     f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                     f'Updating password...')
            cursor.execute(f"UPDATE users SET password='{new_passw}' WHERE name='{active_users[0][1]}' AND "
                           f"account='{active_users[0][2]}' AND phone_number='{active_users[0][4]}';")
            conn.commit()
            log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
                     f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
                     f'Done!')
            return redirect(url_for('admin_profile'))
    else:
        return render_template('admin-update-password.html')


# main function
# noinspection PyShadowingNames
@app.route('/')
def main():
    active_users = []
    cursor.execute("SELECT * FROM users WHERE currently_active='true';")
    for i in cursor.fetchall():
        active_users.append(i)
    cursor.execute("SELECT * FROM users WHERE id=1;")
    if len(active_users) == 0:
        return redirect(url_for('sign_up'))
    if len(active_users) > 0:
        if active_users[0][0] == 1:
            return redirect(url_for('admin_dashboard'))
        if active_users[0][0] != 1:
            return redirect(url_for('dashboard'))


if __name__ == "__main__":
    # set cache limit to 0 (improves performance)
    app.jinja_env.cache = {}
    # run app
    log.info('App successfully started!')
    app.run()
    log.info(f':timestamp(utc): {arrow.utcnow().format("YYYY.MM.DD HH:mm:ss")}:timestamp(sys_time):'
             f'{arrow.utcnow().to(get_localzone()).format("YYYY.MM.DD HH:mm:ss")}; '
             f'App successfully shut down!')

# close mysql connection
conn.close()
